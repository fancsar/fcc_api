# -*- coding: utf-8 -*-
# @Time   : 2019/4/1410:54
# @Author :lemon_FCC
# @Email  :670992243@qq.com
# @File   :do_Excel.py
from openpyxl import load_workbook

# 完成excel的读与写
from API_03.common.do_HTTPrequ import HTTPrequest

'''
执行用例sql时，有可能会用到多条sql
'''
class Case:
    # 测试用例类
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None


class DoExcel:
    def __init__(self, file_name, sheet_name):
        try:  # 异常处理
            self.name = file_name
            self.sheet = sheet_name
            self.workbook = load_workbook(file_name)
        except:
            print('不存在此excel文件')

    # 读取用例
    def get_case(self):
        sheet = self.workbook[self.sheet]
        row = sheet.max_row  # 获取最大行
        cases = []  # 存放所有的用例
        # 遍历要读取的用例数据
        for i in range(2, row + 1):
            case = Case()  # 实例化一个case
            case.case_id = sheet.cell(i, 1).value
            case.title = sheet.cell(i, 2).value
            case.url = sheet.cell(i, 3).value
            case.data = sheet.cell(i, 4).value
            case.method = sheet.cell(i, 5).value
            case.excepted = sheet.cell(i, 6).value
            case.sql = sheet.cell(i, 9).value  # sql
            cases.append(case)
        self.workbook.close()

        return cases

    # 写入用例结果数据
    def write_case(self, row, actual, result):
        sheet = self.workbook[self.sheet]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(self.name)
        self.workbook.close()


#
if __name__ == '__main__':
    from API_02.common.do_contants import case_files
    from API_02.common.do_HTTPrequ import HTTPrequest

    doe = DoExcel(case_files, 'Add')
    cas = doe.get_case()
    for case in cas:
        case.data = eval(case.data)
        print(case.data)
        print(type(case.data))
        if case.data.__contains__('mobilephone') and case.data['mobilephone'] == 'normal_user':
            case.data['mobilephone'] = 13821774377
            print('存在')
        else:
            print('不存在')
            # httpres=HTTPrequest()
            # httpres.
